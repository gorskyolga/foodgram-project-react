import openpyxl
import os
import shutil
from django.contrib.auth import get_user_model
from django.core.management.base import BaseCommand

from recipes.models import Ingredient, Tag, Recipe, TagRecipe, IngredientRecipe

User = get_user_model()


class Command(BaseCommand):
    help = 'Добавление исходных данных в таблицы БД.'

    def handle(self, *args, **options):
        self.stdout.write('Добавление пользователей...')
        wb = openpyxl.load_workbook('/app/data/recipe_data.xlsx')
        sheet = wb['users']
        for row in range(2, sheet.max_row + 1):
            user, created = User.objects.get_or_create(
                email=sheet.cell(row, 4).value,
                username=sheet.cell(row, 3).value,
                first_name=sheet.cell(row, 1).value,
                last_name=sheet.cell(row, 2).value
            )
            user.set_password(sheet.cell(row, 5).value)
            user.save()
        self.stdout.write('Добавление тегов...')
        sheet = wb['tags']
        for row in range(2, sheet.max_row + 1):
            Tag.objects.get_or_create(
                name=sheet.cell(row, 1).value, slug=sheet.cell(row, 2).value,
                color=sheet.cell(row, 3).value
            )
        self.stdout.write('Добавление картинок для рецептов...')
        path_scr = '/app/data/images/'
        path_dst = '/app/media/recipes/images/'
        if not os.path.exists(path_dst):
            os.makedirs(path_dst)
        for file in os.listdir(path_scr):
            shutil.copyfile(path_scr + file, path_dst + file)
        self.stdout.write('Добавление рецептов...')
        sheet = wb['recipes']
        for row in range(2, sheet.max_row + 1):
            author = User.objects.get(username=sheet.cell(row, 1).value)
            Recipe.objects.get_or_create(
                author=author, name=sheet.cell(row, 2).value,
                image=sheet.cell(row, 3).value, text=sheet.cell(row, 4).value,
                cooking_time=sheet.cell(row, 5).value
            )
        self.stdout.write('Добавление связей тегов и рецептов...')
        sheet = wb['recipe_tag']
        for row in range(2, sheet.max_row + 1):
            author = User.objects.get(username=sheet.cell(row, 1).value)
            recipe = Recipe.objects.get(
                author=author, name=sheet.cell(row, 2).value
            )
            tag = Tag.objects.get(slug=sheet.cell(row, 3).value)
            TagRecipe.objects.get_or_create(recipe=recipe, tag=tag)
        self.stdout.write('Добавление связей ингредиентов и рецептов...')
        sheet = wb['recipe_ingredient']
        for row in range(2, sheet.max_row + 1):
            author = User.objects.get(username=sheet.cell(row, 1).value)
            recipe = Recipe.objects.get(
                author=author, name=sheet.cell(row, 2).value
            )
            ingredient = Ingredient.objects.get(
                name=sheet.cell(row, 3).value,
                measurement_unit=sheet.cell(row, 5).value
            )
            IngredientRecipe.objects.get_or_create(
                recipe=recipe, ingredient=ingredient,
                amount=sheet.cell(row, 4).value
            )
        self.stdout.write('Данные загружены.')
